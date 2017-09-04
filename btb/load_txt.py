# -*- coding: utf-8 -*-
from openpyxl import Workbook, load_workbook
import os


def creatwb(xls_name, sheet_name):
    wb = Workbook()
    sheet = wb.active
    sheet.title = sheet_name
    for field in range(1, len(fields) + 1):  # 写入表头
        _ = sheet.cell(row=1, column=field, value=str(fields[field - 1]))
    wb.save(filename=xls_name)
    print("新建Excel：" + xls_name + "成功")


def savetoexcel(data, xls_name):
    print("写入excel：")
    wb = load_workbook(filename=xls_name)
    sheet_worker = wb.active
    for lines in data:
        lines = eval(lines)
        for d in lines:
            sheet_worker.append(d)

    wb.save(filename=xls_name)
    print("保存成功")


if __name__ == '__main__':
    fields = ['成交类型', '成交时间', '成交量	', '成交价格', '总金额', '时间戳']
    judge = os.path.exists('need.xlsx')
    need_xlsx = 'need.xlsx'
    deal_sheet = 'deal_sheet'
    if not judge:
        creatwb(need_xlsx, deal_sheet)
    with open('data.txt', 'r') as f:
        txt_list = f.readlines()
    savetoexcel(txt_list, need_xlsx)
